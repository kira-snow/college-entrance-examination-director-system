# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html
import pymysql
from twisted.enterprise import adbapi
import traceback
import os
import copy
import logging
from scrapy.exceptions import DropItem
from openpyxl import Workbook
from gaokaospider.items import GaokaospiderItem, LinespiderItem, SchoolInfoItem,SegmentsItem
from gaokaospider.settings import LEVEL_DICT,WL_DICT,LEVEL_LIST,WL_LIST,YEARS_START
import queue
import sqlite3

logger = logging.getLogger("GaokaospiderPipeline")

class GaokaospiderPipeline(object):
    def __init__(self, pool, settings):
        self.dbpool = pool
        self.settings = settings
        self.excelFiles_name = []
        self.lastyear = YEARS_START
        excel_path = os.getcwd()
        for level in LEVEL_LIST:
            excel_file = level + '.xlsx'
            self.excelFiles_name.append(os.path.join(excel_path, excel_file))
        self.data_all = []
        for i in range(len(self.excelFiles_name)):
            data_i = []
            data_y = []
            for j in range(len(WL_LIST)):
                data_i.append( {} )
                data_y.append( {} )
            self.data_all.append(data_i)
    
    @classmethod
    def from_settings(cls, settings):
        params = dict(
            # host=settings.get("HOST"),
            # port=settings.get("PORT"),
            db=settings.get("DATABASE") #,
            # user=settings.get("USER"),
            # passwd=settings.get("PASSWORD"),
            # charset=settings.get("CHARSET"),
            # cursorclass=pymysql.cursors.DictCursor
        )
        db_connect_pool = None
        if settings.get("MYSQL"):
            GaokaospiderPipeline.__test_mysql_settings(**params)
            # db_connect_pool = adbapi.ConnectionPool('pymysql', **params)
            db_connect_pool = True
        obj = cls(db_connect_pool, settings)
        return obj

    def open_spider(self, spider):
        self.schoolsQ= queue.Queue(0)
        self.levelLinesQ= queue.Queue(0)
        self.schoolInfoQ= queue.Queue(0)
        self.segmentsQ= queue.Queue(0)
        self.excelFiles = []
        for i in range(len(self.excelFiles_name)):
            self.excelFiles.append( self.newExcelFile() )

    def close_spider(self, spider):
        try:
            for i in range(len(self.excelFiles_name)):
                for ws in self.excelFiles[i]:
                    for j in range(YEARS_START, self.lastyear+1):
                        ws.cell(row=1,column=j-YEARS_START+3,value=j)
                self.excelFiles[i].save(self.excelFiles_name[i])
                logger.info("excel文件已存储于 " + self.excelFiles_name[i])
            if self.dbpool:
                # self.dbpool.close()
                sqlite_conn = sqlite3.connect(self.settings.get("DATABASE")+".db")
                sqlite_cursor = sqlite_conn.cursor()
                sqlite_conn.execute("BEGIN TRANSACTION;")
                while not self.schoolsQ.empty():
                    try:
                        sqlite_cursor.execute(self.schoolsQ.get())
                    except Exception as e:
                        logger.error(traceback.format_exc())
                        continue
                sqlite_conn.execute("COMMIT;")

                sqlite_conn.execute("BEGIN TRANSACTION;")
                while not self.levelLinesQ.empty():
                    try:
                        sql_insert = self.levelLinesQ.get()
                        sqlite_cursor.execute(sql_insert)
                    except Exception as e:
                        logger.error(sql_insert + ' ERROR:' +traceback.format_exc())
                        continue
                sqlite_conn.execute("COMMIT;")

                sqlite_conn.execute("BEGIN TRANSACTION;")
                while not self.schoolInfoQ.empty():
                    try:
                        sqlite_cursor.execute(self.schoolInfoQ.get())
                    except Exception as e:
                        logger.error(traceback.format_exc())
                        continue
                sqlite_conn.execute("COMMIT;")

                sqlite_conn.execute("BEGIN TRANSACTION;")
                while not self.segmentsQ.empty():
                    try:
                        sqlite_cursor.execute(self.segmentsQ.get())
                    except Exception as e:
                        logger.error(traceback.format_exc())
                        continue
                sqlite_conn.execute("COMMIT;")
                sqlite_conn.close()
                logger.info("数据已存储于数据库" + self.settings.get("DATABASE") + ".db")
        except Exception as e:
            logger.error(traceback.format_exc())

    def process_item(self, item, spider):
        if isinstance(item, GaokaospiderItem):
            self.process_item_gaokao(item, spider)
        elif isinstance(item, LinespiderItem):
            self.process_item_line(item, spider)
        elif isinstance(item, SchoolInfoItem):
            self.process_item_schoolInfo(item, spider)
        elif isinstance(item, SegmentsItem):
            self.process_item_segment(item, spider)
    
    def process_item_gaokao(self, item, spider):
        try:
            wb = self.excelFiles[item['level']]
            work_sheet = wb.get_sheet_by_name( WL_LIST[item['wl']] )
            # work_sheet.append([item['id'], item['school'], item['score']])
            col = item['year'] - YEARS_START + 3
            if item['year'] > self.lastyear:
                self.lastyear = item['year']
            lv = item['level']
            wl = item['wl']
            sch = item['school']
            sch=sch.strip()
            if sch == '':
                print('------------空sch：'+LEVEL_LIST[item['level']]+','+WL_LIST[item['wl']]+','+item['id']+','+str(item['year']) )
                return item
            if sch in self.data_all[lv][wl].keys():
                rw = self.data_all[lv][wl][sch]
                work_sheet.cell(row=rw,column=col,value=item['score'])
                work_sheet.cell(row=rw,column=1,value=item['id'])
            else:
                work_sheet.append([item['id'], sch])
                work_sheet.cell(row=work_sheet.max_row,column=col,value=item['score'])
                self.data_all[lv][wl][sch] = work_sheet.max_row
            if self.dbpool:
                mysql_item=copy.deepcopy(item)
                mysql_item['school'] = sch
                mysql_item['level'] = LEVEL_LIST[lv]
                mysql_item['wl'] = WL_LIST[wl]
                # result = self.dbpool.runInteraction(self.insert, mysql_item)
                self.insert(mysql_item)
                # 给result绑定一个回调函数，用于监听错误信息
                #result.addErrback(self.error)
            return item
        except Exception as e:
            logger.critical(traceback.format_exc())
            raise DropItem(traceback.format_exc())

    def process_item_line(self, item, spider):
        try:
            if self.dbpool:
                mysql_item=copy.deepcopy(item)
                if mysql_item['level'] == '1':
                    mysql_item1=copy.deepcopy(mysql_item)
                    mysql_item1['level']='1a'
                    # result = self.dbpool.runInteraction(self.insert_line, mysql_item1)
                    self.insert_line(mysql_item1)
                    # 给result绑定一个回调函数，用于监听错误信息
                    #result.addErrback(self.error)
                    mysql_item2=copy.deepcopy(mysql_item)
                    mysql_item2['level']='1a1'
                    # result = self.dbpool.runInteraction(self.insert_line, mysql_item2)
                    self.insert_line(mysql_item2)
                    # 给result绑定一个回调函数，用于监听错误信息
                    #result.addErrback(self.error)
                    mysql_item3=copy.deepcopy(mysql_item)
                    mysql_item3['level']='1b'
                    # result = self.dbpool.runInteraction(self.insert_line, mysql_item3)
                    self.insert_line(mysql_item3)
                    # 给result绑定一个回调函数，用于监听错误信息
                    #result.addErrback(self.error)
                elif mysql_item['level'] == '2':
                    mysql_item1=copy.deepcopy(mysql_item)
                    mysql_item1['level']='2a'
                    # result = self.dbpool.runInteraction(self.insert_line, mysql_item1)
                    self.insert_line(mysql_item1)
                    # 给result绑定一个回调函数，用于监听错误信息
                    #result.addErrback(self.error)
                    mysql_item2=copy.deepcopy(mysql_item)
                    mysql_item2['level']='2b'
                    # result = self.dbpool.runInteraction(self.insert_line, mysql_item2)
                    self.insert_line(mysql_item2)
                    # 给result绑定一个回调函数，用于监听错误信息
                    #result.addErrback(self.error)
                else:
                    # result = self.dbpool.runInteraction(self.insert_line, mysql_item)
                    self.insert_line(mysql_item)
                    # 给result绑定一个回调函数，用于监听错误信息
                    #result.addErrback(self.error)
            return item
        except Exception as e:
            logger.critical(traceback.format_exc())
            raise DropItem(traceback.format_exc())

    def process_item_schoolInfo(self, item, spider):
        try:
            if self.dbpool:
                mysql_item=copy.deepcopy(item)
                # result = self.dbpool.runInteraction(self.insert_schoolInfo, mysql_item)
                self.insert_schoolInfo(mysql_item)
                # 给result绑定一个回调函数，用于监听错误信息
                #result.addErrback(self.error)
            return item
        except Exception as e:
            logger.critical(traceback.format_exc())
            raise DropItem(traceback.format_exc())

    def process_item_segment(self, item, spider):
        try:
            if self.dbpool:
                mysql_item=copy.deepcopy(item)
                # result = self.dbpool.runInteraction(self.insert_segment, mysql_item)
                self.insert_segment(mysql_item)
                # 给result绑定一个回调函数，用于监听错误信息
                #result.addErrback(self.error)
            return item
        except Exception as e:
            logger.critical(traceback.format_exc())
            raise DropItem(traceback.format_exc())

    def insert_schoolInfo(self, item): # (self, cursor, item):
        insert_sql = "insert into schoolInfo (`name`, `province`, `department`, `type`, `level`, `yldx`, `ylxk`, `yjsy`) VALUES (" \
            "'{0}','{1}','{2}','{3}','{4}','{5}','{6}','{7}'); " \
            .format(item['name'], item['province'], item['department'], item['type'], item['level'], item['yldx'], item['ylxk'], item['yjsy'])
        # cursor.execute(insert_sql)
        # logger.info("schoolInfoQ: " +insert_sql)
        self.schoolInfoQ.put(insert_sql)

    def insert_line(self, item): #(self, cursor, item)
        insert_sql = "insert into levelLines (`year`, `level`, `wl`, `type`, `score`) VALUES (" \
            "'{0}','{1}','{2}','{3}',{4}); " \
            .format(item['year'], item['level'], item['wl'], item['type'], item['score'])
        # cursor.execute(insert_sql)
        # logger.info("levelLinesQ: " +insert_sql)
        self.levelLinesQ.put(insert_sql)

    def insert(self, item): #(self, cursor, item)
        insert_sql = "insert into schools (`schoolId`, `name`, `year`, `level`, `wl`, `score`) VALUES (" \
            "'{0}','{1}','{2}','{3}','{4}',{5}); " \
            .format(item['id'], item['school'], item['year'], item['level'], item['wl'], item['score'])
        # cursor.execute(insert_sql)
        # logger.info("schoolsQ: " +insert_sql)
        self.schoolsQ.put(insert_sql)

    def insert_segment(self, item): #(self, cursor, item)
        insert_sql = "insert into segments (`year`, `wl`, `score`, `ranking`) VALUES (" \
            "'{0}','{1}',{2},{3}); " \
            .format(item['year'], item['wl'], item['score'], item['ranking'])
        print(insert_sql)
        # cursor.execute(insert_sql)
        # logger.info("segmentsQ: " +insert_sql)
        self.segmentsQ.put(insert_sql)

    def error(self, reason):
        # 跳过主键重复error
        logger.error("insert to database err: -------------\n" + reason.getErrorMessage() + "\n" + str(reason.getTraceback()))

    def newExcelFile(self):
        wbk = Workbook()
        for i in range(len(WL_LIST)):
            work_sheet = wbk.create_sheet(WL_LIST[i], i)
            work_sheet.append(['院校代码', '院校名称'])
            work_sheet.column_dimensions['B'].width = 50.0
        return wbk


    @staticmethod
    def __test_mysql_settings(**params):
        try:
            # db = pymysql.connect(**params)
            # db.close()
            sqlite_conn = sqlite3.connect(params["db"]+".db")
            sqlite_cursor = sqlite_conn.cursor()
            sql='''
            CREATE TABLE IF NOT EXISTS levellines (
            year varchar(10) NOT NULL,
            level varchar(40) NOT NULL,
            wl varchar(10) NOT NULL,
            type varchar(10) NOT NULL,
            score float NOT NULL,
            PRIMARY KEY (year,level,wl,type)
            ) ;
            '''
            logger.info("sqlite: " +sql)
            sqlite_cursor.execute(sql)
            sql='''
            CREATE TABLE IF NOT EXISTS schoolinfo (
            name varchar(100) NOT NULL,
            province varchar(100) NOT NULL,
            department varchar(100) NOT NULL,
            type varchar(10) NOT NULL,
            level varchar(10) NOT NULL,
            yldx char(1) DEFAULT NULL,
            ylxk char(1) DEFAULT NULL,
            yjsy char(1) DEFAULT NULL,
            PRIMARY KEY (name)
            );
            '''
            logger.info("sqlite: " +sql)
            sqlite_cursor.execute(sql)
            sql='''
            CREATE TABLE IF NOT EXISTS schools (
            schoolId varchar(10) NOT NULL,
            name varchar(100) NOT NULL,
            year varchar(10) NOT NULL,
            level varchar(40) NOT NULL,
            wl varchar(10) NOT NULL,
            score float NOT NULL,
            PRIMARY KEY (schoolId,year,level,wl)
            ) ;
            '''
            logger.info("sqlite: " +sql)
            sqlite_cursor.execute(sql)
            sql='''
            CREATE TABLE IF NOT EXISTS segments (
            year varchar(10) NOT NULL,
            wl varchar(10) NOT NULL,
            score float NOT NULL,
            ranking int NOT NULL,
            PRIMARY KEY (year,wl,score)
            ) ;
            '''
            logger.info("sqlite: " +sql)
            sqlite_cursor.execute(sql)
            sqlite_conn.commit()
            sqlite_conn.close()
        except Exception as e:
            logger.critical(str(e))
            os._exit(1)